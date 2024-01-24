


import requests
from bs4 import BeautifulSoup
import pandas as pd





#scraped text list
global_list = []




def scrape_single_page(url):
    response = requests.get(url)
    res = ""
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        res = soup.find('div', class_="td-post-content tagdiv-type")
        if res is None:
            res = soup.find('div',class_="tdb-block-inner td-fix-index")
        return res.get_text()
    else:
        print(f"Failed to fetch {url}")





def main():

    df = pd.read_excel("input.xlsx")
    url_list = df['URL']
    for url in url_list:
        data = scrape_single_page(url)
        global_list.append(data)





import nltk
from nltk.corpus import stopwords

def remove_stop_words(sentences):
    nltk.download('stopwords')
    stop_words = set(stopwords.words('english'))
    new_sentences = [' '.join([word for word in sentence.split() if word not in stop_words]) for sentence in sentences]
    return new_sentences

print(remove_stop_words(global_list))





main()
new_data_scrapping = remove_stop_words(global_list)

#total number of words in each sentence of list
total_word_after_cleaning = []
for sentence in new_data_scrapping:
    len_sen = len(sentence)
    total_word_after_cleaning.append(len_sen)

# total_word_after_cleaning
# len(total_word_after_cleaning)
new_data_scrapping





#positive word list
my_file = open("positive_words.txt", "r", encoding='utf-8')
data = my_file.read()
positive_word_list = data.split("\n")
my_file.close()




def positive_score(data_scrapping, positive_words):
    pos_score = []
    for sentence in data_scrapping:
        score = 0
        for word in sentence.split():
            if word in positive_words:
                score = score+1
        pos_score.append(score)
    return pos_score





pos_score_list = positive_score(new_data_scrapping,positive_word_list)
len(pos_score_list)





#negative word list
my_file = open("negative_words.txt", "r", encoding='utf-8')
data = my_file.read()
negative_word_list = data.split("\n")
my_file.close()





def negative_score(data_scrapping, negative_words):
    neg_score = []
    for sentence in data_scrapping:
        score = 0
        for word in sentence.split():
            if word in negative_words:
                score = score+1
        neg_score.append(score)

    return neg_score





neg_score_list = negative_score(new_data_scrapping,negative_word_list)
len(neg_score_list)





def polarity_score(positive_score, negative_score):
    polarity = []
    for i in range(len(positive_score)):
        polar = (positive_score[i] - negative_score[i])/((positive_score[i] + negative_score[i])+0.000001)
        polarity.append(polar)

    return polarity





pol_score_list = polarity_score(pos_score_list,neg_score_list)
len(pol_score_list)





#subjective score
def subjective_score(positive_score, negative_score,total_word_after_cleaning):
    subjective_score = []
    for i in range(len(positive_score)):
        sub = (positive_score[i] + negative_score[i])/((total_word_after_cleaning[i]) + 0.000001)
        subjective_score.append(sub)
    return subjective_score





sub_score_list = subjective_score(pos_score_list, neg_score_list, total_word_after_cleaning)
len(sub_score_list)





#average_sentance length

def avgg(data):
    avg_len = []
    for string in data:
        l = 0
        x = string.split('.')
        total_words = len(x)
        for sentence in x:
            l+=len(sentence.split())
        avgr = l/total_words

        avg_len.append(avgr)

    return avg_len






avg_sent_len = avgg(new_data_scrapping)
len(avg_sent_len)





# percentage of complex words

my_file = open("complex_word_list.txt", "r", encoding='utf-8')
data = my_file.read()
complex_word_list = data.split("\n")
my_file.close()

comp_score_list = []
for sentence in new_data_scrapping:
    score = 0
    for word in sentence:
        if word in complex_word_list:
            score = score+1
    comp_score_list.append(score)

per_comp_score_list = []

for x in range(len(comp_score_list)):
    x=sum(comp_score_list)/sum(total_word_after_cleaning)
    per_comp_score_list.append(x)

per_comp_sc_l = per_comp_score_list





len(per_comp_sc_l)




# fog index

fog_index = []

for i in range(len(avg_sent_len)):
    f = 0.4*(avg_sent_len[i] + per_comp_score_list[i])
    fog_index.append(f)

fog_index
len(fog_index)





# avg no of words per sentence

avg_no_words_sent = avg_sent_len
avg_no_words_sent
len(avg_no_words_sent)





#complex word_count

comp_word_count = comp_score_list
comp_word_count
len(comp_word_count)





# word count

word_count = total_word_after_cleaning
len(word_count)





# syllable count per word
import re

def count_syllables(word):
    word = re.sub(r'(es|ed)$', '', word, flags=re.IGNORECASE)

    vowels = re.findall(r'[aeiou]', word, flags=re.IGNORECASE)

    syllable_count = len(re.findall(r'[aeiou]+', ''.join(vowels), flags=re.IGNORECASE))

    return syllable_count

def count_syllables_per_word(sentence):
    words = sentence.split()
    syllable_counts = [count_syllables(word) for word in words]
    return syllable_counts

def count_syllables_in_sentences(sentences):
    result = []
    for sentence in sentences:
        syllable_counts = count_syllables_per_word(sentence)
        result.append(syllable_counts)
    return result

syllable_counts_per_sentence = count_syllables_in_sentences(new_data_scrapping)

syll_count = []
for i, counts in enumerate(syllable_counts_per_sentence):
    syll_count.append(sum(counts))

syll_count
len(syll_count)





import re

def count_personal_pronouns(sentences):
    personal_pronouns_regex = r'\b(?:I|we|my|ours|us)\b'
    count = []

    for sentence in sentences:
        c = 0
        for words in sentence.split():
            matches = re.findall(personal_pronouns_regex, sentence, flags=re.IGNORECASE)
            c += 1
        count.append(c)

    return count

pronoun_count = count_personal_pronouns(new_data_scrapping)
pronoun_count

len(pronoun_count)





# average word length

def average_word_length(datas):
    avg_word_len = []
    for words in datas:
        each_data_len = 0
        for word in words.split():
            each_data_len+=len(word)

        if len(words) == 0:
            avg_word_len.append(0)
        else:
            x = int(each_data_len/len(words.split()))
            avg_word_len.append(x)

    return avg_word_len

average_word_length = average_word_length(new_data_scrapping)
average_word_length
len(average_word_length)




import openpyxl


workbook = openpyxl.load_workbook('output_data_new_ex.xlsx')
sheet = workbook['Sheet1']


# positive score
data_list_c = pos_score_list
column_letter_c = 'C'
next_row_c = sheet.max_row + 1
for value in data_list_c:
    sheet[column_letter_c + str(next_row_c)] = value
    next_row_c += 1



# negative score
data_list_d = neg_score_list
column_letter_d = 'D'
next_row_d = sheet.max_row + 1
for value in data_list_d:
    sheet[column_letter_d + str(next_row_d)] = value
    next_row_d += 1



# polarity score
data_list_e = pol_score_list
column_letter_e = 'E'
next_row_e = sheet.max_row + 1
for value in data_list_e:
    sheet[column_letter_e + str(next_row_e)] = value
    next_row_e += 1



# subjective score
data_list_f = sub_score_list
column_letter_f = 'F'
next_row_f = sheet.max_row + 1
for value in data_list_f:
    sheet[column_letter_f + str(next_row_f)] = value
    next_row_f += 1


#average sentence length
data_list_g = avg_sent_len
column_letter_g = 'G'
next_row_g = sheet.max_row + 1
for value in data_list_g:
    sheet[column_letter_g + str(next_row_g)] = value
    next_row_g += 1


# percentage of complex words
data_list_h = per_comp_score_list
column_letter_h = 'H'
next_row_h = sheet.max_row + 1
for value in data_list_h:
    sheet[column_letter_h + str(next_row_h)] = value
    next_row_h += 1


# fog index
data_list_i = fog_index
column_letter_i = 'I'
next_row_i = sheet.max_row + 1
for value in data_list_i:
    sheet[column_letter_i + str(next_row_i)] = value
    next_row_i += 1


#average number of words per sentence
data_list_j = avg_no_words_sent
column_letter_j = 'J'
next_row_j = sheet.max_row + 1
for value in data_list_j:
    sheet[column_letter_j + str(next_row_j)] = value
    next_row_j += 1


# complex word count
data_list_k = comp_word_count
column_letter_k = 'K'
next_row_k = sheet.max_row + 1
for value in data_list_k:
    sheet[column_letter_k + str(next_row_k)] = value
    next_row_k += 1


# word count
data_list_l = word_count
column_letter_l = 'L'
next_row_l = sheet.max_row + 1
for value in data_list_l:
    sheet[column_letter_l + str(next_row_l)] = value
    next_row_l += 1


# syllable per word
data_list_m = syll_count
column_letter_m = 'M'
next_row_m = sheet.max_row + 1
for value in data_list_m:
    sheet[column_letter_m + str(next_row_m)] = value
    next_row_m += 1


# personal pronoun
data_list_n = pronoun_count
column_letter_n = 'N'
next_row_n = sheet.max_row + 1
for value in data_list_n:
    sheet[column_letter_n + str(next_row_n)] = value
    next_row_n += 1


# average word length
data_list_o = average_word_length
column_letter_o = 'O'
next_row_o = sheet.max_row + 1
for value in data_list_o:
    sheet[column_letter_o + str(next_row_o)] = value
    next_row_o += 1



workbook.save('new_output_file.xlsx')








